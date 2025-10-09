#!/usr/bin/env python3
"""
Lead Enrichment System

This script:
1. Reads leads from Excel file
2. Ingests them into Cosmos DB
3. Scrapes SAM.gov project pages for additional details
4. Enriches Cosmos DB records with scraped data
5. Qualifies leads against company database

Usage:
    python lead_enrichment.py --file "Top Leads Review Phase1_10032025 (2).xlsx" --action ingest
    python lead_enrichment.py --action scrape
    python lead_enrichment.py --action qualify
"""

import os
import sys
import json
import time
import argparse
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

# Azure Cosmos DB
from azure.cosmos import CosmosClient, PartitionKey, exceptions

# Load environment
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Cosmos DB Configuration
COSMOS_ENDPOINT = os.getenv("COSMOS_ENDPOINT")
COSMOS_KEY = os.getenv("COSMOS_KEY")
DATABASE_NAME = "DefianceDB"
LEADS_CONTAINER = "Leads"
COMPANIES_CONTAINER = "wbi_info"  # Your company database


class LeadEnrichmentSystem:
    """Main system for lead enrichment"""

    def __init__(self):
        """Initialize Cosmos DB clients"""
        self.cosmos_client = CosmosClient(COSMOS_ENDPOINT, COSMOS_KEY)
        self.database = self.cosmos_client.get_database_client(DATABASE_NAME)

        # Create Leads container if it doesn't exist
        try:
            self.leads_container = self.database.create_container(
                id=LEADS_CONTAINER,
                partition_key=PartitionKey(path="/id"),
                offer_throughput=400
            )
            logger.info(f"Created container: {LEADS_CONTAINER}")
        except exceptions.CosmosResourceExistsError:
            self.leads_container = self.database.get_container_client(LEADS_CONTAINER)
            logger.info(f"Using existing container: {LEADS_CONTAINER}")

        # Company database for qualification
        try:
            self.companies_container = self.database.get_container_client(COMPANIES_CONTAINER)
            logger.info(f"Connected to company database: {COMPANIES_CONTAINER}")
        except Exception as e:
            logger.warning(f"Could not connect to company database: {e}")
            self.companies_container = None

    def read_excel_leads(self, file_path: str) -> List[Dict[str, Any]]:
        """Read leads from Excel file"""
        logger.info(f"Reading Excel file: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Get headers
        headers = [cell.value for cell in sheet[1]]

        leads = []
        for row_idx in range(2, sheet.max_row + 1):
            lead = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value is not None:
                    lead[header] = cell_value

            if lead:  # Only add non-empty rows
                # Generate unique ID
                lead['id'] = f"lead_{row_idx - 1}_{int(time.time())}"
                lead['ingested_at'] = datetime.utcnow().isoformat()
                lead['enrichment_status'] = 'pending'
                lead['scraped_data'] = {}
                lead['qualification_status'] = 'pending'
                lead['qualification_score'] = None

                leads.append(lead)

        logger.info(f"Read {len(leads)} leads from Excel")
        return leads

    def ingest_leads(self, leads: List[Dict[str, Any]]) -> int:
        """Ingest leads into Cosmos DB"""
        logger.info(f"Ingesting {len(leads)} leads into Cosmos DB...")

        ingested_count = 0
        for lead in leads:
            try:
                self.leads_container.upsert_item(lead)
                ingested_count += 1
                logger.info(f"Ingested: {lead.get('Opportunity Title', 'Unknown')[:50]}")
            except Exception as e:
                logger.error(f"Failed to ingest lead {lead['id']}: {e}")

        logger.info(f"Successfully ingested {ingested_count}/{len(leads)} leads")
        return ingested_count

    def scrape_sam_gov_page(self, url: str) -> Dict[str, Any]:
        """
        Scrape SAM.gov opportunity page for additional information

        Note: SAM.gov has bot protection. This is a basic scraper that may need
        to be enhanced with Selenium/Playwright for JS-rendered content.
        """
        logger.info(f"Scraping: {url}")

        scraped_data = {
            'url': url,
            'scraped_at': datetime.utcnow().isoformat(),
            'scrape_status': 'pending'
        }

        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }

            response = requests.get(url, headers=headers, timeout=30)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, 'html.parser')

            # Extract common SAM.gov fields
            # Note: These selectors may need adjustment based on actual page structure

            # Try to extract solicitation number
            solicitation = soup.find('span', {'id': 'solicitationNumber'})
            if solicitation:
                scraped_data['solicitation_number'] = solicitation.text.strip()

            # Try to extract agency
            agency = soup.find('span', {'id': 'agency'})
            if agency:
                scraped_data['agency'] = agency.text.strip()

            # Try to extract description
            description = soup.find('div', {'id': 'description'})
            if description:
                scraped_data['description'] = description.text.strip()

            # Try to extract deadlines
            deadline = soup.find('span', {'id': 'responseDeadline'})
            if deadline:
                scraped_data['response_deadline'] = deadline.text.strip()

            # Try to extract NAICS code
            naics = soup.find('span', {'id': 'naicsCode'})
            if naics:
                scraped_data['naics_code'] = naics.text.strip()

            # Try to extract set-aside info
            set_aside = soup.find('span', {'id': 'setAside'})
            if set_aside:
                scraped_data['set_aside'] = set_aside.text.strip()

            # Try to extract place of performance
            place = soup.find('div', {'id': 'placeOfPerformance'})
            if place:
                scraped_data['place_of_performance'] = place.text.strip()

            # Extract all text for general content
            scraped_data['full_text'] = soup.get_text(separator=' ', strip=True)[:10000]  # Limit to 10K chars

            scraped_data['scrape_status'] = 'success'
            logger.info(f"Successfully scraped: {url}")

        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to scrape {url}: {e}")
            scraped_data['scrape_status'] = 'failed'
            scraped_data['scrape_error'] = str(e)
        except Exception as e:
            logger.error(f"Unexpected error scraping {url}: {e}")
            scraped_data['scrape_status'] = 'failed'
            scraped_data['scrape_error'] = str(e)

        return scraped_data

    def enrich_leads_with_scraping(self, batch_size: int = 10) -> int:
        """
        Scrape and enrich leads that have 'Link' field and pending enrichment status
        """
        logger.info("Starting lead enrichment with web scraping...")

        # Query for leads with pending enrichment
        query = "SELECT * FROM c WHERE c.enrichment_status = 'pending' AND IS_DEFINED(c.Link)"

        leads_to_enrich = list(self.leads_container.query_items(
            query=query,
            enable_cross_partition_query=True
        ))

        logger.info(f"Found {len(leads_to_enrich)} leads to enrich")

        enriched_count = 0
        for i, lead in enumerate(leads_to_enrich[:batch_size]):
            logger.info(f"Enriching lead {i+1}/{min(batch_size, len(leads_to_enrich))}: {lead['Opportunity Title'][:50]}")

            url = lead.get('Link')
            if not url:
                continue

            # Scrape the page
            scraped_data = self.scrape_sam_gov_page(url)

            # Update lead with scraped data
            lead['scraped_data'] = scraped_data
            lead['enrichment_status'] = 'completed' if scraped_data['scrape_status'] == 'success' else 'failed'
            lead['enriched_at'] = datetime.utcnow().isoformat()

            try:
                self.leads_container.upsert_item(lead)
                enriched_count += 1
                logger.info(f"Enriched lead: {lead['Opportunity Title'][:50]}")
            except Exception as e:
                logger.error(f"Failed to update lead {lead['id']}: {e}")

            # Rate limiting - be respectful to SAM.gov
            time.sleep(2)

        logger.info(f"Enriched {enriched_count} leads")
        return enriched_count

    def qualify_leads(self) -> int:
        """
        Qualify leads based on company database information

        This compares lead data against your company capabilities in wbi_info container
        """
        if not self.companies_container:
            logger.error("Company database not available for qualification")
            return 0

        logger.info("Starting lead qualification...")

        # Query for enriched leads with pending qualification
        query = "SELECT * FROM c WHERE c.enrichment_status = 'completed' AND c.qualification_status = 'pending'"

        leads_to_qualify = list(self.leads_container.query_items(
            query=query,
            enable_cross_partition_query=True
        ))

        logger.info(f"Found {len(leads_to_qualify)} leads to qualify")

        # Get company capabilities
        company_query = "SELECT * FROM c"
        company_data = list(self.companies_container.query_items(
            query=company_query,
            enable_cross_partition_query=True
        ))

        logger.info(f"Loaded {len(company_data)} company records")

        qualified_count = 0
        for lead in leads_to_qualify:
            # Simple qualification logic - you can enhance this
            qualification_score = 0.0
            qualification_reasons = []

            # Check if recommended offering matches company capabilities
            recommended_offering = lead.get('Recommended WBI Offering', '').lower()

            # Check scraped data for relevant keywords
            scraped_text = lead.get('scraped_data', {}).get('full_text', '').lower()

            # Match against company data (customize based on your wbi_info structure)
            for company_record in company_data:
                company_text = json.dumps(company_record).lower()

                # Simple keyword matching - enhance with NLP/embeddings
                if recommended_offering and recommended_offering in company_text:
                    qualification_score += 0.3
                    qualification_reasons.append(f"Offering '{recommended_offering}' matches company capabilities")

                # Check for technology keywords
                tech_keywords = ['ai', 'machine learning', 'data analytics', 'cloud', 'automation']
                for keyword in tech_keywords:
                    if keyword in scraped_text and keyword in company_text:
                        qualification_score += 0.1
                        qualification_reasons.append(f"Technology match: {keyword}")

            # Normalize score
            qualification_score = min(1.0, qualification_score)

            # Update lead with qualification
            lead['qualification_status'] = 'completed'
            lead['qualification_score'] = qualification_score
            lead['qualification_reasons'] = qualification_reasons
            lead['qualified_at'] = datetime.utcnow().isoformat()

            try:
                self.leads_container.upsert_item(lead)
                qualified_count += 1
                logger.info(f"Qualified: {lead['Opportunity Title'][:50]} - Score: {qualification_score:.2f}")
            except Exception as e:
                logger.error(f"Failed to update lead {lead['id']}: {e}")

        logger.info(f"Qualified {qualified_count} leads")
        return qualified_count

    def get_lead_statistics(self) -> Dict[str, Any]:
        """Get statistics about leads in the database"""
        logger.info("Calculating lead statistics...")

        stats = {
            'total_leads': 0,
            'enrichment_pending': 0,
            'enrichment_completed': 0,
            'enrichment_failed': 0,
            'qualification_pending': 0,
            'qualification_completed': 0,
            'average_qualification_score': 0.0,
            'high_quality_leads': 0  # Score > 0.7
        }

        # Get all leads
        all_leads = list(self.leads_container.query_items(
            query="SELECT * FROM c",
            enable_cross_partition_query=True
        ))

        stats['total_leads'] = len(all_leads)

        qualification_scores = []
        for lead in all_leads:
            # Enrichment stats
            if lead.get('enrichment_status') == 'pending':
                stats['enrichment_pending'] += 1
            elif lead.get('enrichment_status') == 'completed':
                stats['enrichment_completed'] += 1
            elif lead.get('enrichment_status') == 'failed':
                stats['enrichment_failed'] += 1

            # Qualification stats
            if lead.get('qualification_status') == 'pending':
                stats['qualification_pending'] += 1
            elif lead.get('qualification_status') == 'completed':
                stats['qualification_completed'] += 1
                score = lead.get('qualification_score', 0)
                qualification_scores.append(score)
                if score > 0.7:
                    stats['high_quality_leads'] += 1

        if qualification_scores:
            stats['average_qualification_score'] = sum(qualification_scores) / len(qualification_scores)

        return stats


def main():
    """Main CLI entry point"""
    parser = argparse.ArgumentParser(description='Lead Enrichment System')
    parser.add_argument('--file', type=str, help='Path to Excel file with leads')
    parser.add_argument('--action', type=str, required=True,
                       choices=['ingest', 'scrape', 'qualify', 'stats', 'full'],
                       help='Action to perform')
    parser.add_argument('--batch-size', type=int, default=10,
                       help='Number of leads to process in batch')

    args = parser.parse_args()

    # Initialize system
    system = LeadEnrichmentSystem()

    if args.action == 'ingest':
        if not args.file:
            logger.error("--file required for ingest action")
            sys.exit(1)

        leads = system.read_excel_leads(args.file)
        system.ingest_leads(leads)

    elif args.action == 'scrape':
        system.enrich_leads_with_scraping(batch_size=args.batch_size)

    elif args.action == 'qualify':
        system.qualify_leads()

    elif args.action == 'stats':
        stats = system.get_lead_statistics()
        print("\n=== Lead Statistics ===")
        for key, value in stats.items():
            print(f"{key}: {value}")

    elif args.action == 'full':
        # Run full pipeline
        if not args.file:
            logger.error("--file required for full pipeline")
            sys.exit(1)

        logger.info("Running full pipeline...")
        leads = system.read_excel_leads(args.file)
        system.ingest_leads(leads)
        system.enrich_leads_with_scraping(batch_size=args.batch_size)
        system.qualify_leads()

        stats = system.get_lead_statistics()
        print("\n=== Final Statistics ===")
        for key, value in stats.items():
            print(f"{key}: {value}")


if __name__ == "__main__":
    main()
