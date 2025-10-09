#!/usr/bin/env python3
"""
Adaptive Lead Enrichment System

Strategy: "If there's a link, go gather information"

This system:
1. Automatically detects ANY links in Excel/CSV files (no hardcoded columns)
2. Scrapes those links intelligently
3. Uses LLM to extract structured information
4. Enriches Cosmos DB with discovered data
5. Qualifies leads against company database using AI

Usage:
    # Ingest any file with links
    python adaptive_lead_enrichment.py --file "your_file.xlsx" --action ingest

    # Scrape all pending links
    python adaptive_lead_enrichment.py --action scrape --batch 5

    # Qualify leads using AI
    python adaptive_lead_enrichment.py --action qualify

    # Full pipeline
    python adaptive_lead_enrichment.py --file "your_file.xlsx" --action full --batch 10
"""

import os
import re
import sys
import json
import time
import argparse
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import uuid

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openai import AzureOpenAI

# Azure Cosmos DB
from azure.cosmos import CosmosClient, PartitionKey, exceptions

# Load environment
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Cosmos DB Configuration
COSMOS_ENDPOINT = os.getenv("COSMOS_ENDPOINT")
COSMOS_KEY = os.getenv("COSMOS_KEY")
DATABASE_NAME = "DefianceDB"
LEADS_CONTAINER = "Leads"
COMPANIES_CONTAINER = "wbi_info"

# Azure OpenAI for intelligent extraction
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_KEY")
GPT41_DEPLOYMENT = os.getenv("GPT41_DEPLOYMENT", "gpt-4")


class AdaptiveLeadEnrichmentSystem:
    """Adaptive system that handles ANY file format with links"""

    def __init__(self):
        """Initialize clients"""
        self.cosmos_client = CosmosClient(COSMOS_ENDPOINT, COSMOS_KEY)
        self.database = self.cosmos_client.get_database_client(DATABASE_NAME)

        # Initialize OpenAI client
        self.openai_client = AzureOpenAI(
            azure_endpoint=AZURE_OPENAI_ENDPOINT,
            api_key=AZURE_OPENAI_KEY,
            api_version="2024-02-15-preview"
        )

        # Create/get Leads container
        try:
            self.leads_container = self.database.create_container(
                id=LEADS_CONTAINER,
                partition_key=PartitionKey(path="/id")
                # Note: No offer_throughput for serverless accounts
            )
            logger.info(f"Created container: {LEADS_CONTAINER}")
        except exceptions.CosmosResourceExistsError:
            self.leads_container = self.database.get_container_client(LEADS_CONTAINER)
            logger.info(f"Using existing container: {LEADS_CONTAINER}")

        # Company database
        try:
            self.companies_container = self.database.get_container_client(COMPANIES_CONTAINER)
            logger.info(f"Connected to company database: {COMPANIES_CONTAINER}")
        except Exception as e:
            logger.warning(f"Could not connect to company database: {e}")
            self.companies_container = None

    def detect_links_in_row(self, row_data: Dict[str, Any]) -> List[Tuple[str, str]]:
        """
        Intelligently detect ALL links in a row of data

        Returns: List of (field_name, url) tuples
        """
        url_pattern = re.compile(
            r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
        )

        links = []
        for field_name, value in row_data.items():
            if not value:
                continue

            value_str = str(value)

            # Check if entire value is a URL
            if url_pattern.match(value_str):
                links.append((field_name, value_str))
            # Check if value contains URLs
            elif url_pattern.search(value_str):
                found_urls = url_pattern.findall(value_str)
                for url in found_urls:
                    links.append((field_name, url))

        return links

    def read_file_adaptively(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Adaptively read ANY file format (Excel, CSV, etc.)
        Auto-detect structure and extract links
        """
        logger.info(f"Adaptively reading file: {file_path}")

        file_ext = Path(file_path).suffix.lower()

        # Read based on file type
        # For Excel, try openpyxl directly for better handling of complex files
        if file_ext in ['.xlsx', '.xls']:
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    headers.append(cell.value if cell.value else f"Unnamed_{cell.column}")

                # Read data rows
                data = []
                for row_idx in range(2, sheet.max_row + 1):
                    row_dict = {}
                    for col_idx, header in enumerate(headers, 1):
                        cell_value = sheet.cell(row_idx, col_idx).value
                        row_dict[header] = cell_value
                    data.append(row_dict)

                df = pd.DataFrame(data)
                logger.info(f"Read {len(df)} rows using openpyxl")

            except Exception as e:
                logger.warning(f"Openpyxl failed, trying pandas: {e}")
                df = pd.read_excel(file_path)

        elif file_ext == '.csv':
            df = pd.read_csv(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")

        logger.info(f"Read {len(df)} rows with columns: {list(df.columns)}")

        leads = []
        for idx, row in df.iterrows():
            # Convert row to dict
            row_dict = row.to_dict()

            # Detect all links in this row
            detected_links = self.detect_links_in_row(row_dict)

            if not detected_links:
                logger.debug(f"Row {idx}: No links found, skipping")
                continue

            # Create lead entry
            lead = {
                'id': f"lead_{uuid.uuid4()}",
                'ingested_at': datetime.utcnow().isoformat(),
                'source_file': Path(file_path).name,
                'source_row': int(idx) + 2,  # Excel row number (1-indexed + header)
                'original_data': row_dict,
                'detected_links': [
                    {'field': field, 'url': url} for field, url in detected_links
                ],
                'enrichment_status': 'pending',
                'enriched_data': {},
                'qualification_status': 'pending',
                'qualification_result': {}
            }

            leads.append(lead)
            logger.info(f"Row {idx+2}: Found {len(detected_links)} link(s)")

        logger.info(f"Extracted {len(leads)} leads with links from file")
        return leads

    def mark_link_as_scraped(self, lead_id: str, url: str, status: str, error: str = None):
        """
        Mark a specific link as scraped in the database for resume capability
        """
        try:
            lead = self.leads_container.read_item(item=lead_id, partition_key=lead_id)

            if 'scrape_history' not in lead:
                lead['scrape_history'] = {}

            lead['scrape_history'][url] = {
                'status': status,
                'scraped_at': datetime.utcnow().isoformat(),
                'error': error
            }

            self.leads_container.upsert_item(lead)
            logger.info(f"Marked {url} as {status}")
        except Exception as e:
            logger.error(f"Failed to mark link as scraped: {e}")

    def is_link_already_scraped(self, lead: Dict[str, Any], url: str) -> bool:
        """
        Check if a link has already been successfully scraped
        """
        scrape_history = lead.get('scrape_history', {})
        return url in scrape_history and scrape_history[url].get('status') == 'success'

    def scrape_url_intelligently(self, url: str, lead_id: str = None) -> Dict[str, Any]:
        """
        Intelligently scrape ANY URL and extract relevant information
        Marks link as scraped in database for resume capability
        """
        logger.info(f"Scraping: {url}")

        result = {
            'url': url,
            'scraped_at': datetime.utcnow().isoformat(),
            'status': 'pending',
            'raw_html_length': 0,
            'extracted_text': '',
            'structured_data': {}
        }

        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }

            response = requests.get(url, headers=headers, timeout=30, allow_redirects=True)
            response.raise_for_status()

            result['final_url'] = response.url
            result['raw_html_length'] = len(response.content)

            soup = BeautifulSoup(response.content, 'html.parser')

            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()

            # Get text
            text = soup.get_text(separator='\n', strip=True)
            result['extracted_text'] = text[:50000]  # Limit to 50K chars

            # Extract metadata
            result['structured_data']['title'] = soup.title.string if soup.title else None

            # Try to extract common metadata
            meta_tags = {}
            for meta in soup.find_all('meta'):
                name = meta.get('name') or meta.get('property')
                content = meta.get('content')
                if name and content:
                    meta_tags[name] = content

            result['structured_data']['meta_tags'] = meta_tags

            # Extract all links
            links = [a.get('href') for a in soup.find_all('a', href=True)]
            result['structured_data']['all_links'] = links[:100]  # Limit to 100

            result['status'] = 'success'
            logger.info(f"Successfully scraped {url} - {len(text)} chars of text")

            # Mark as successfully scraped in database
            if lead_id:
                self.mark_link_as_scraped(lead_id, url, 'success')

        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to scrape {url}: {e}")
            result['status'] = 'failed'
            result['error'] = str(e)

            # Mark as failed in database
            if lead_id:
                self.mark_link_as_scraped(lead_id, url, 'failed', str(e))

        except Exception as e:
            logger.error(f"Unexpected error scraping {url}: {e}")
            result['status'] = 'failed'
            result['error'] = str(e)

            # Mark as failed in database
            if lead_id:
                self.mark_link_as_scraped(lead_id, url, 'failed', str(e))

        return result

    def extract_structured_info_with_llm(self, scraped_data: Dict[str, Any], original_context: Dict[str, Any]) -> Dict[str, Any]:
        """
        Use LLM to intelligently extract structured information from scraped content
        """
        if scraped_data.get('status') != 'success':
            return {'extraction_status': 'skipped', 'reason': 'Scrape failed'}

        text = scraped_data.get('extracted_text', '')
        if not text or len(text) < 100:
            return {'extraction_status': 'skipped', 'reason': 'Insufficient text'}

        # Truncate text for LLM
        text_sample = text[:15000]  # Use first 15K chars

        prompt = f"""You are an expert at extracting structured information from web pages.

ORIGINAL LEAD CONTEXT:
{json.dumps(original_context, indent=2)}

SCRAPED WEB PAGE TEXT:
{text_sample}

Extract ALL relevant information in JSON format. Be comprehensive - include:
1. Project/opportunity details (title, description, objectives)
2. Organization information (agency, company, department)
3. Requirements (technical, business, compliance)
4. Timeline information (deadlines, milestones, duration)
5. Financial information (budget, contract value, funding)
6. Contact information (emails, names, roles)
7. Location information (place of performance, addresses)
8. Classification (NAICS codes, categories, types)
9. Key themes and topics
10. Any other relevant structured data

Respond ONLY with valid JSON containing the extracted information. Use nested structures where appropriate."""

        try:
            logger.info("Extracting structured data with LLM...")

            response = self.openai_client.chat.completions.create(
                model=GPT41_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": "You are an expert data extraction assistant. Always respond with valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.1
            )

            extracted = json.loads(response.choices[0].message.content)
            extracted['extraction_status'] = 'success'
            extracted['extracted_at'] = datetime.utcnow().isoformat()

            logger.info("Successfully extracted structured data with LLM")
            return extracted

        except Exception as e:
            logger.error(f"LLM extraction failed: {e}")
            return {
                'extraction_status': 'failed',
                'error': str(e)
            }

    def enrich_leads_with_scraping(self, batch_size: int = 5, resume: bool = True) -> int:
        """
        Enrich leads by scraping all detected links and using LLM for extraction

        Args:
            batch_size: Number of leads to process
            resume: If True, skip links that were already successfully scraped

        Returns:
            Number of leads enriched

        Resume Logic:
            - Each link is marked as scraped in database immediately after scraping
            - If process crashes, restart with --action scrape --resume
            - Already scraped links will be skipped automatically
        """
        logger.info(f"Starting adaptive lead enrichment (resume={resume})...")

        # Query for pending OR in-progress leads (for resume capability)
        query = "SELECT * FROM c WHERE c.enrichment_status IN ('pending', 'in_progress')"

        pending_leads = list(self.leads_container.query_items(
            query=query,
            enable_cross_partition_query=True
        ))

        logger.info(f"Found {len(pending_leads)} leads to process")

        enriched_count = 0
        for i, lead in enumerate(pending_leads[:batch_size]):
            lead_id = lead['id']

            logger.info(f"\n{'='*60}")
            logger.info(f"Enriching lead {i+1}/{min(batch_size, len(pending_leads))}")
            logger.info(f"Lead ID: {lead_id}")
            logger.info(f"Source: {lead.get('source_file')} Row {lead.get('source_row')}")

            # Mark lead as in_progress
            lead['enrichment_status'] = 'in_progress'
            lead['enrichment_started_at'] = datetime.utcnow().isoformat()
            try:
                self.leads_container.upsert_item(lead)
            except Exception as e:
                logger.error(f"Failed to mark lead as in_progress: {e}")
                continue

            all_scraped_data = []
            links_to_scrape = lead.get('detected_links', [])
            total_links = len(links_to_scrape)

            # Scrape each detected link
            for link_idx, link_info in enumerate(links_to_scrape, 1):
                url = link_info['url']
                field = link_info['field']

                # Check if already scraped (resume capability)
                if resume and self.is_link_already_scraped(lead, url):
                    logger.info(f"[{link_idx}/{total_links}] ✓ SKIPPING (already scraped): {url}")

                    # Still need to include in scraped_data for completeness
                    scrape_record = lead.get('scrape_history', {}).get(url, {})
                    all_scraped_data.append({
                        'source_field': field,
                        'url': url,
                        'scraped_data': {
                            'url': url,
                            'status': 'previously_scraped',
                            'scraped_at': scrape_record.get('scraped_at')
                        },
                        'structured_extraction': {'extraction_status': 'skipped', 'reason': 'Previously scraped'}
                    })
                    continue

                logger.info(f"[{link_idx}/{total_links}] Scraping link from field '{field}': {url}")

                try:
                    # Scrape the URL (will mark as scraped in database)
                    scraped = self.scrape_url_intelligently(url, lead_id=lead_id)

                    # Extract structured info with LLM (only if scrape succeeded)
                    if scraped['status'] == 'success':
                        structured = self.extract_structured_info_with_llm(
                            scraped,
                            lead.get('original_data', {})
                        )
                    else:
                        structured = {'extraction_status': 'skipped', 'reason': 'Scrape failed'}

                    all_scraped_data.append({
                        'source_field': field,
                        'url': url,
                        'scraped_data': scraped,
                        'structured_extraction': structured
                    })

                    # Rate limiting
                    time.sleep(3)

                except Exception as e:
                    logger.error(f"Error processing link {url}: {e}")
                    # Mark as failed and continue
                    self.mark_link_as_scraped(lead_id, url, 'failed', str(e))
                    all_scraped_data.append({
                        'source_field': field,
                        'url': url,
                        'scraped_data': {'status': 'failed', 'error': str(e)},
                        'structured_extraction': {'extraction_status': 'failed'}
                    })
                    continue

            # Update lead
            lead['enriched_data'] = {
                'scraped_links': all_scraped_data,
                'total_links_scraped': len(all_scraped_data),
                'successful_scrapes': sum(1 for s in all_scraped_data if s['scraped_data']['status'] == 'success')
            }
            lead['enrichment_status'] = 'completed'
            lead['enriched_at'] = datetime.utcnow().isoformat()

            try:
                self.leads_container.upsert_item(lead)
                enriched_count += 1
                logger.info(f"✅ Enriched lead {lead['id']}")
            except Exception as e:
                logger.error(f"Failed to update lead: {e}")

        logger.info(f"\n{'='*60}")
        logger.info(f"Enrichment complete: {enriched_count}/{min(batch_size, len(pending_leads))} leads")
        return enriched_count

    def qualify_lead_with_ai(self, lead: Dict[str, Any], company_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Use AI to qualify a lead against company capabilities
        """
        logger.info("Qualifying lead with AI...")

        # Build context
        lead_summary = {
            'original_data': lead.get('original_data', {}),
            'enriched_data_summary': {
                'total_links_scraped': lead.get('enriched_data', {}).get('total_links_scraped', 0),
                'extracted_info': []
            }
        }

        # Collect key extracted info
        for scraped_link in lead.get('enriched_data', {}).get('scraped_links', []):
            extraction = scraped_link.get('structured_extraction', {})
            if extraction.get('extraction_status') == 'success':
                lead_summary['enriched_data_summary']['extracted_info'].append(extraction)

        # Build company summary
        company_summary = []
        for company_record in company_data[:10]:  # Limit to 10 records for context
            company_summary.append(company_record)

        prompt = f"""You are an expert business development analyst. Analyze this lead and qualify it against company capabilities.

LEAD INFORMATION:
{json.dumps(lead_summary, indent=2)}

COMPANY CAPABILITIES (Sample):
{json.dumps(company_summary, indent=2)}

Analyze and respond with JSON containing:
{{
  "qualification_score": 0.0-1.0,
  "recommendation": "pursue|review|pass",
  "strengths": ["list of alignment points"],
  "concerns": ["list of potential issues"],
  "required_capabilities": ["capabilities needed to win"],
  "company_capability_matches": ["specific company capabilities that align"],
  "strategic_value": "high|medium|low",
  "reasoning": "detailed explanation of qualification decision"
}}

Be thorough and analytical."""

        try:
            response = self.openai_client.chat.completions.create(
                model=GPT41_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": "You are an expert business development analyst with deep experience in government contracting and capability assessment."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.2
            )

            qualification = json.loads(response.choices[0].message.content)
            qualification['qualified_at'] = datetime.utcnow().isoformat()
            qualification['qualification_method'] = 'ai_analysis'

            logger.info(f"Qualification complete - Score: {qualification.get('qualification_score', 0):.2f}, Recommendation: {qualification.get('recommendation')}")
            return qualification

        except Exception as e:
            logger.error(f"AI qualification failed: {e}")
            return {
                'qualification_score': 0.0,
                'recommendation': 'review',
                'error': str(e),
                'qualification_method': 'failed'
            }

    def qualify_leads(self, batch_size: int = 10) -> int:
        """
        Qualify enriched leads using AI
        """
        if not self.companies_container:
            logger.error("Company database not available")
            return 0

        logger.info("Starting AI-powered lead qualification...")

        # Get enriched leads pending qualification
        query = "SELECT * FROM c WHERE c.enrichment_status = 'completed' AND c.qualification_status = 'pending'"

        leads_to_qualify = list(self.leads_container.query_items(
            query=query,
            enable_cross_partition_query=True
        ))

        logger.info(f"Found {len(leads_to_qualify)} leads to qualify")

        # Load company data
        company_query = "SELECT * FROM c"
        company_data = list(self.companies_container.query_items(
            query=company_query,
            enable_cross_partition_query=True
        ))

        logger.info(f"Loaded {len(company_data)} company records")

        qualified_count = 0
        for i, lead in enumerate(leads_to_qualify[:batch_size]):
            logger.info(f"\nQualifying lead {i+1}/{min(batch_size, len(leads_to_qualify))}")

            # Qualify with AI
            qualification = self.qualify_lead_with_ai(lead, company_data)

            # Update lead
            lead['qualification_result'] = qualification
            lead['qualification_status'] = 'completed'

            try:
                self.leads_container.upsert_item(lead)
                qualified_count += 1
                logger.info(f"✅ Qualified lead {lead['id']}")
            except Exception as e:
                logger.error(f"Failed to update lead: {e}")

            # Rate limiting
            time.sleep(2)

        logger.info(f"\nQualification complete: {qualified_count}/{min(batch_size, len(leads_to_qualify))} leads")
        return qualified_count

    def get_statistics(self) -> Dict[str, Any]:
        """Get comprehensive statistics"""
        all_leads = list(self.leads_container.query_items(
            query="SELECT * FROM c",
            enable_cross_partition_query=True
        ))

        stats = {
            'total_leads': len(all_leads),
            'by_enrichment_status': {},
            'by_qualification_status': {},
            'qualification_scores': {
                'high': 0,  # > 0.7
                'medium': 0,  # 0.4-0.7
                'low': 0  # < 0.4
            },
            'recommendations': {
                'pursue': 0,
                'review': 0,
                'pass': 0
            }
        }

        for lead in all_leads:
            # Enrichment status
            enrich_status = lead.get('enrichment_status', 'unknown')
            stats['by_enrichment_status'][enrich_status] = stats['by_enrichment_status'].get(enrich_status, 0) + 1

            # Qualification status
            qual_status = lead.get('qualification_status', 'unknown')
            stats['by_qualification_status'][qual_status] = stats['by_qualification_status'].get(qual_status, 0) + 1

            # Qualification scores
            qual_result = lead.get('qualification_result', {})
            score = qual_result.get('qualification_score', 0)
            if score > 0.7:
                stats['qualification_scores']['high'] += 1
            elif score >= 0.4:
                stats['qualification_scores']['medium'] += 1
            elif score > 0:
                stats['qualification_scores']['low'] += 1

            # Recommendations
            recommendation = qual_result.get('recommendation', 'unknown')
            if recommendation in stats['recommendations']:
                stats['recommendations'][recommendation] += 1

        return stats


def main():
    """CLI entry point"""
    parser = argparse.ArgumentParser(description='Adaptive Lead Enrichment System')
    parser.add_argument('--file', type=str, help='Path to file with leads (Excel/CSV)')
    parser.add_argument('--action', type=str, required=True,
                       choices=['ingest', 'scrape', 'qualify', 'stats', 'full'],
                       help='Action to perform')
    parser.add_argument('--batch', type=int, default=5,
                       help='Batch size for processing')
    parser.add_argument('--resume', action='store_true',
                       help='Resume from last checkpoint (skip already scraped links)')

    args = parser.parse_args()

    # Initialize system
    system = AdaptiveLeadEnrichmentSystem()

    if args.action == 'ingest':
        if not args.file:
            logger.error("--file required for ingest")
            sys.exit(1)

        leads = system.read_file_adaptively(args.file)

        # Ingest to Cosmos DB
        ingested = 0
        for lead in leads:
            try:
                system.leads_container.upsert_item(lead)
                ingested += 1
            except Exception as e:
                logger.error(f"Failed to ingest lead: {e}")

        logger.info(f"Ingested {ingested}/{len(leads)} leads")

    elif args.action == 'scrape':
        system.enrich_leads_with_scraping(batch_size=args.batch, resume=args.resume)

    elif args.action == 'qualify':
        system.qualify_leads(batch_size=args.batch)

    elif args.action == 'stats':
        stats = system.get_statistics()
        print("\n" + "="*60)
        print("LEAD STATISTICS")
        print("="*60)
        print(json.dumps(stats, indent=2))

    elif args.action == 'full':
        if not args.file:
            logger.error("--file required for full pipeline")
            sys.exit(1)

        logger.info("\n" + "="*60)
        logger.info("RUNNING FULL PIPELINE")
        logger.info("="*60 + "\n")

        # Ingest
        leads = system.read_file_adaptively(args.file)
        for lead in leads:
            try:
                system.leads_container.upsert_item(lead)
            except Exception as e:
                logger.error(f"Failed to ingest: {e}")

        # Scrape & enrich
        system.enrich_leads_with_scraping(batch_size=args.batch, resume=args.resume)

        # Qualify
        system.qualify_leads(batch_size=args.batch)

        # Stats
        stats = system.get_statistics()
        print("\n" + "="*60)
        print("FINAL STATISTICS")
        print("="*60)
        print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
